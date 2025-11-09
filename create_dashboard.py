import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

print("Starting dashboard creation...")

# --- 1. Create Workbook and Sheets ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Dashboard"

# Create Data sheets
ws_data = wb.create_sheet("Data")
ws_data.append(["timestamp", "metric_key", "current_value", "previous_value", "metric_label"])
ws_data.freeze_panes = "A2"

ws_history = wb.create_sheet("Data History")
ws_history.append(["sync_date", "metric_key", "value", "period", "notes"])
ws_history.freeze_panes = "A2"


# --- 2. Define Styles (with Montserrat) ---
font_title = Font(name="Montserrat", size=24, bold=True, color="FFFFFF")
font_header = Font(name="Montserrat", size=13, bold=True, color="FFFFFF")
font_metric_value = Font(name="Montserrat", size=20, bold=True, color="2D323C")
font_metric_label = Font(name="Montserrat", size=9, color="A4ACBA")
font_trend = Font(name="Montserrat", size=9, color="808080")
font_secondary = Font(name="Montserrat", size=10, color="2D323C")
font_content = Font(name="Montserrat", size=9, color="A4ACBA")
font_control = Font(name="Montserrat", size=10)
font_updated = Font(name="Montserrat", size=10, color="808080")

align_center_middle = Alignment(horizontal="center", vertical="center", wrap_text=False)
align_center_bottom = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
align_right_middle = Alignment(horizontal="right", vertical="center")

fill_title = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
fill_gsc = PatternFill(start_color="1C4C8A", end_color="1C4C8A", fill_type="solid")
fill_ig = PatternFill(start_color="E4405F", end_color="E4405F", fill_type="solid")
fill_fb = PatternFill(start_color="1877F2", end_color="1877F2", fill_type="solid")
fill_yt = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
fill_li = PatternFill(start_color="0A66C2", end_color="0A66C2", fill_type="solid")
fill_mc = PatternFill(start_color="FFE01B", end_color="FFE01B", fill_type="solid")
font_mc_header = Font(name="Montserrat", size=13, bold=True, color="000000") # Black text for yellow bg

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
card_outline = Border(
    left=Side(style='medium', color='BFBFBF'),
    right=Side(style='medium', color='BFBFBF'),
    top=Side(style='medium', color='BFBFBF'),
    bottom=Side(style='medium', color='BFBFBF')
)

# --- 3. Setup Dashboard Layout ---

# Column widths
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 3
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 3
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 15
ws.column_dimensions['L'].width = 15

# Title
ws.merge_cells('B2:L2')
ws['B2'] = "MARKETING DASHBOARD"
ws['B2'].font = font_title
ws['B2'].fill = fill_title
ws['B2'].alignment = align_center_middle
ws.row_dimensions[2].height = 45

# Controls
ws.row_dimensions[3].height = 30
ws['B3'].font = font_control
ws['K3'] = "Last Updated:"
ws['K3'].font = font_updated
ws['K3'].alignment = align_right_middle
ws['L3'].font = font_updated
ws['L3'].alignment = align_center_middle # Script will populate this

# Time period dropdown
dv = DataValidation(type="list", formula1='"Last 7 Days,Last 30 Days,Last 60 Days,Last 90 Days"', allow_blank=True)
dv.error = "Invalid selection"
dv.errorTitle = "Invalid Time Period"
dv.prompt = "Please select a time period"
dv.promptTitle = "Time Period"
ws.add_data_validation(dv)
dv.add('B3')
ws['B3'] = "Last 30 Days"

# Row heights for cards
ws.row_dimensions[4].height = 8
ws.row_dimensions[13].height = 8
ws.row_dimensions[22].height = 8

card_rows = [5, 14] # Start rows for card rows 1, 2
for row_start in card_rows:
    ws.row_dimensions[row_start].height = 30     # Header
    ws.row_dimensions[row_start + 1].height = 8  # Spacer
    ws.row_dimensions[row_start + 2].height = 18 # Labels
    ws.row_dimensions[row_start + 3].height = 32 # Values
    ws.row_dimensions[row_start + 4].height = 18 # Trends
    ws.row_dimensions[row_start + 5].height = 8  # Spacer
    ws.row_dimensions[row_start + 6].height = 22 # Secondary
    ws.row_dimensions[row_start + 7].height = 20 # Content
    
# --- 4. Define Card Generator ---

def create_card(ws, config):
    start_row = config['start_row']
    start_col_letter = config['start_col']
    start_col_idx = openpyxl.utils.column_index_from_string(start_col_letter)
    
    end_col_letter = chr(ord(start_col_letter) + 2)
    merge_range = f"{start_col_letter}{start_row}:{end_col_letter}{start_row}"
    
    # Card Header
    ws.merge_cells(merge_range)
    header_cell = ws[f"{start_col_letter}{start_row}"]
    header_cell.value = f"{config['emoji']} {config['name']}"
    header_cell.fill = config['fill']
    header_cell.font = config.get('font', font_header)
    header_cell.alignment = align_center_middle
    
    # Apply outline to card
    for row in range(start_row, start_row + 8):
        for col_idx in range(start_col_idx, start_col_idx + 3):
            cell = ws.cell(row=row, column=col_idx)
            # Apply thin inner borders
            if row > start_row:
                cell.border = thin_border
            
    # Apply thick outer border
    for row in range(start_row, start_row + 8):
        ws.cell(row=row, column=start_col_idx).border = Border(left=card_outline.left, top=thin_border.top, bottom=thin_border.bottom, right=thin_border.right)
        ws.cell(row=row, column=start_col_idx+2).border = Border(right=card_outline.right, top=thin_border.top, bottom=thin_border.bottom, left=thin_border.left)
    for col_idx in range(start_col_idx, start_col_idx + 3):
        ws.cell(row=start_row, column=col_idx).border = Border(top=card_outline.top, left=thin_border.left, right=thin_border.right) # Header gets top border
        ws.cell(row=start_row+7, column=col_idx).border = Border(bottom=card_outline.bottom, left=thin_border.left, right=thin_border.right)


    # Primary Metrics (Labels, Values, Trends)
    for i, metric in enumerate(config['metrics']):
        col_letter = chr(ord(start_col_letter) + i)
        
        # Label
        label_cell = ws[f"{col_letter}{start_row + 2}"]
        label_cell.value = f'=IFERROR(INDEX(Data!E:E,MATCH("{metric["key"]}",Data!B:B,0)),"{metric["label"]}")'
        label_cell.font = font_metric_label
        label_cell.alignment = align_center_bottom
        
        # Value
        value_cell = ws[f"{col_letter}{start_row + 3}"]
        base_formula = f'IFERROR(VLOOKUP("{metric["key"]}",Data!B:C,2,FALSE),0)'
        
        if metric['format'] == 'number':
            value_cell.value = f'=LET(val,{base_formula}, IF(val>=1000000,TEXT(val/1000000,"0.0")&"M", IF(val>=1000,TEXT(val/1000,"0.0")&"K", TEXT(val,"#,##0"))))'
        elif metric['format'] == 'decimal':
             value_cell.value = f'=LET(val,{base_formula}, IF(val>=1000000,TEXT(val/1000000,"0.0")&"M", IF(val>=1000,TEXT(val/1000,"0.0")&"K", TEXT(val,"#,##0.0"))))'
        elif metric['format'] == 'percent':
            value_cell.value = f'=TEXT({base_formula},"0.00%")'
            
        value_cell.font = font_metric_value
        value_cell.alignment = align_center_middle
        
        # Trend
        trend_cell = ws[f"{col_letter}{start_row + 4}"]
        curr_formula = f'IFERROR(VLOOKUP("{metric["key"]}",Data!B:C,2,FALSE),0)'
        prev_formula = f'IFERROR(VLOOKUP("{metric["key"]}",Data!B:D,3,FALSE),0)'
        trend_cell.value = f'=LET(curr,{curr_formula}, prev,{prev_formula}, IF(prev=0,"—",TEXT((curr-prev)/prev,"↑0.0%;↓0.0%;—")) )'
        trend_cell.font = font_trend
        trend_cell.alignment = align_center_middle

    # Secondary & Content
    secondary_range = f"{start_col_letter}{start_row + 6}:{end_col_letter}{start_row + 6}"
    ws.merge_cells(secondary_range)
    secondary_cell = ws[f"{start_col_letter}{start_row + 6}"]
    secondary_cell.value = config['secondary']
    secondary_cell.font = font_secondary
    secondary_cell.alignment = align_center_middle
    
    content_range = f"{start_col_letter}{start_row + 7}:{end_col_letter}{start_row + 7}"
    ws.merge_cells(content_range)
    content_cell = ws[f"{start_col_letter}{start_row + 7}"]
    content_cell.value = config['content']
    content_cell.font = font_content
    content_cell.alignment = align_center_middle

# --- 5. Define All Cards ---

card_definitions = [
    {
        'name': "GOOGLE SEARCH CONSOLE", 'emoji': "🔍", 'fill': fill_gsc,
        'start_row': 5, 'start_col': 'B',
        'metrics': [
            {'key': 'gsc_clicks', 'label': 'Clicks', 'format': 'number'},
            {'key': 'gsc_impressions', 'label': 'Impressions', 'format': 'number'},
            {'key': 'gsc_ctr', 'label': 'Avg. CTR', 'format': 'percent'},
        ],
        'secondary': '="Avg. " & TEXT(IFERROR(VLOOKUP("gsc_position",Data!B:C,2,FALSE),0),"0.0") & " position"',
        'content': ""
    },
    {
        'name': "INSTAGRAM", 'emoji': "📸", 'fill': fill_ig,
        'start_row': 5, 'start_col': 'F',
        'metrics': [
            {'key': 'ig_followers', 'label': 'Followers', 'format': 'number'},
            {'key': 'ig_reach', 'label': 'Reach', 'format': 'number'},
            {'key': 'ig_engagement', 'label': 'Engagement', 'format': 'number'},
        ],
        'secondary': '="Eng. Rate " & TEXT(IFERROR(VLOOKUP("ig_engagement",Data!B:C,2,FALSE)/VLOOKUP("ig_reach",Data!B:C,2,FALSE),0),"0.0%") & " • " & LET(val,IFERROR(VLOOKUP("ig_saves",Data!B:C,2,FALSE),0), IF(val>=1000,TEXT(val/1000,"0.0")&"K", TEXT(val,"#,##0"))) & " saves"',
        'content': '=IFERROR(VLOOKUP("ig_posts",Data!B:C,2,FALSE),0) & " posts • " & IFERROR(VLOOKUP("ig_stories",Data!B:C,2,FALSE),0) & " stories • " & IFERROR(VLOOKUP("ig_reels",Data!B:C,2,FALSE),0) & " reels"'
    },
    {
        'name': "FACEBOOK PAGE", 'emoji': "👍", 'fill': fill_fb,
        'start_row': 5, 'start_col': 'J',
        'metrics': [
            {'key': 'fb_followers', 'label': 'Followers', 'format': 'number'},
            {'key': 'fb_reach', 'label': 'Reach', 'format': 'number'},
            {'key': 'fb_engagement', 'label': 'Engagement', 'format': 'number'},
        ],
        'secondary': '="Eng. Rate " & TEXT(IFERROR(VLOOKUP("fb_engagement",Data!B:C,2,FALSE)/VLOOKUP("fb_reach",Data!B:C,2,FALSE),0),"0.0%") & " • " & LET(val,IFERROR(VLOOKUP("fb_video_views",Data!B:C,2,FALSE),0), IF(val>=1000,TEXT(val/1000,"0.0")&"K", TEXT(val,"#,##0"))) & " video views"',
        'content': '=IFERROR(VLOOKUP("fb_posts",Data!B:C,2,FALSE),0) & " posts • " & IFERROR(VLOOKUP("fb_videos",Data!B:C,2,FALSE),0) & " videos"'
    },
    {
        'name': "YOUTUBE CHANNEL", 'emoji': "▶️", 'fill': fill_yt,
        'start_row': 14, 'start_col': 'B',
        'metrics': [
            {'key': 'yt_subscribers', 'label': 'Subscribers', 'format': 'number'},
            {'key': 'yt_views', 'label': 'Views', 'format': 'number'},
            {'key': 'yt_watch_time', 'label': 'Watch Time (hrs)', 'format': 'decimal'},
        ],
        'secondary': '=LET(val,IFERROR(VLOOKUP("yt_impressions",Data!B:C,2,FALSE),0), IF(val>=1000,TEXT(val/1000,"0.0")&"K", TEXT(val,"#,##0"))) & " impr • " & TEXT(IFERROR(VLOOKUP("yt_ctr",Data!B:C,2,FALSE),0),"0.0%") & " CTR"',
        'content': '=IFERROR(VLOOKUP("yt_videos",Data!B:C,2,FALSE),0) & " videos published"'
    },
    {
        'name': "LINKEDIN PROFILE", 'emoji': "💼", 'fill': fill_li,
        'start_row': 14, 'start_col': 'F',
        'metrics': [
            {'key': 'li_followers', 'label': 'Followers', 'format': 'number'},
            {'key': 'li_impressions', 'label': 'Impressions', 'format': 'number'},
            {'key': 'li_engagement', 'label': 'Engagement', 'format': 'number'},
        ],
        'secondary': '="Eng. Rate " & TEXT(IFERROR(VLOOKUP("li_engagement",Data!B:C,2,FALSE)/VLOOKUP("li_impressions",Data!B:C,2,FALSE),0),"0.0%") & " • " & LET(val,IFERROR(VLOOKUP("li_clicks",Data!B:C,2,FALSE),0), IF(val>=1000,TEXT(val/1000,"0.0")&"K", TEXT(val,"#,##0"))) & " clicks"',
        'content': '=IFERROR(VLOOKUP("li_posts",Data!B:C,2,FALSE),0) & " posts published"'
    },
    {
        'name': "MAILCHIMP", 'emoji': "🐵", 'fill': fill_mc,
        'font': font_mc_header,
        'start_row': 14, 'start_col': 'J',
        'metrics': [
            {'key': 'mc_list_size', 'label': 'List Size', 'format': 'number'},
            {'key': 'mc_open_rate', 'label': 'Open Rate', 'format': 'percent'},
            {'key': 'mc_click_rate', 'label': 'Click Rate', 'format': 'percent'},
        ],
        'secondary': '=LET(val,IFERROR(VLOOKUP("mc_emails_sent",Data!B:C,2,FALSE),0), IF(val>=1000,TEXT(val/1000,"0.0")&"K", TEXT(val,"#,##0"))) & " emails sent"',
        'content': '=IFERROR(VLOOKUP("mc_campaigns",Data!B:C,2,FALSE),0) & " campaigns sent"'
    },
]

# --- 6. Build All Cards ---
for card_config in card_definitions:
    create_card(ws, card_config)

# --- 7. Final Protection and Save ---
ws.protection.sheet = True
ws.protection.enable()

file_name = "Marketing_Dashboard.xlsx"
wb.save(file_name)

print(f"✅ Success! File '{file_name}' has been created.")